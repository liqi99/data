import tushare as ts

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import column_index_from_string
import os
import pandas as pd
from datetime import datetime
import sys
from ReadIn import readOneCsv

pd.set_option('display.max_rows',None)

from Polish import amount

import akshare as ak


def polish(x):
    alist = [x.b1_p,x.b1_v,x.b2_p,x.b2_v,x.b3_p,x.b3_v,x.b4_p,x.b4_v,x.b5_p,x.b5_v]
    blist = [x.a1_p,x.a1_v,x.a2_p,x.a2_v,x.a3_p,x.a3_v,x.a4_p,x.a4_v,x.a5_p,x.a5_v]
    #blist.reverse()
    oline = [alist[2*i]+':'+alist[2*i+1] for i in range(1)]
    oline = ', '.join(oline)

    sline = [blist[2*i]+':'+blist[2*i+1] for i in range(1)]
    sline = ', '.join(sline)
    
    return oline,sline


def high_low(x):
    last = float(x.pre_close)
    high = float(x.high)
    low = float(x.low)
    high = round((high/last-1)*100,2)
    low = round((low/last-1)*100,2)
    high = str(high)+'%'
    low = str(low)+'%'
    return high+','+low
    

def price(x):
    last = float(x.pre_close)
    price = float(x.price)
    price = round((price/last-1)*100,2)
    price = str(price)+'%'
    return price


def get_bj_stock(alist):
    adf = ak.stock_zh_a_spot_em()
    #adf['代码'] = adf['代码'].astype(str)
    adf = adf[adf['代码'].isin(alist)]
    adf['成交量'] = adf['成交量'].apply(lambda x: amount(x))
    adf['总市值'] = adf['总市值'].apply(lambda x: amount(x))
    adf['流通市值'] = adf['流通市值'].apply(lambda x: amount(x))
    adf['成交额'] = adf['成交额'].apply(lambda x: amount(x))
    adf['涨跌幅'] = adf['涨跌幅'].apply(lambda x: str(x)+'%')
    adf['振幅'] = adf['振幅'].apply(lambda x: str(x)+'%')
    #obj_codes = ['838971','833346']
    names = ['代码','名称','最新价','涨跌幅','涨跌额', '市净率','总市值','流通市值','涨速','5分钟涨跌','成交额','成交量']#,'振幅']#,'换手率']
    adf = adf[names]
    print(adf)

def get_stock_data(code, save_file='000001.xlsx'):
    data = ts.get_realtime_quotes(code)
    data['amount'] = data['amount'].apply(lambda x: amount(x))
    data['volume'] = data['volume'].apply(lambda x: str(round(int(x)/1000000,2))+'万')
    data['买盘'] = data.apply(lambda x: polish(x)[0],axis=1)
    data['卖盘'] = data.apply(lambda x: polish(x)[1],axis=1)
    data['最高最低'] = data.apply(lambda x: high_low(x),axis=1)
    data['涨幅'] = data.apply(lambda x: price(x),axis=1)
    data['order'] = data['涨幅'].apply(lambda x: float(x.replace('%','')))
    data = data.sort_values(by=['order'],ascending=False).reset_index(drop=True)

    #data = data.sort_values(by=['price'],ascending=False)
    names = {'volume':'成交量','amount':'成交额','pre_close':'昨收','high':'最高','low':'最低','open':'今开','price':'市价','bid':'买入价',
             'ask':'卖出价'}
    data = data.rename(columns=names)
    #print(data.columns)
    #print(data)

    print_name = ['name','code','date','今开','最高最低','市价','涨幅','买盘','卖盘','成交量','成交额']
    data = data[print_name]
    print(data)


    #if len(sys.argv) == 1:
    #    os.system('date')
    #    exit()
    #else:
    #    date = datetime.today().strftime('%Y-%m-%d')
    #    os.system('python getStockPrice.py {}'.format(sys.argv[1]))
    ##adf = pd.read_csv("./dump/{}/stock_data_{}.csv".format(date[:4],date))
    ##print(adf[['股票代码','股票名称','涨跌幅']])

    #os.system('date')
    #exit()

    ##wb = openpyxl.Workbook()
    ##ws = wb.active

    ## 设置表头
    #header = ['股票代码', '名称', '最新价', '涨跌幅', '成交量', '成交额','盘口']
    #for i in range(len(header)):
    #    cell = ws.cell(row=1, column=i+1)
    #    cell.value = header[i]
    #    print(cell.value)
    #    cell.font = Font(bold=True)
    #    cell.alignment = Alignment(horizontal='center')

    #print(data)
    #exit()

    ## 填充数据
    #row_num = 2
    #for i in range(len(data)):
    #    cell = ws.cell(row=row_num+i, column=column_index_from_string('A'))
    #    cell.value = data.iloc[i]['code']

    #    cell = ws.cell(row=row_num+i, column=column_index_from_string('B'))
    #    cell.value = data.iloc[i]['name']

    #    cell = ws.cell(row=row_num+i, column=column_index_from_string('C'))
    #    cell.value = float(data.iloc[i]['price'])

    #    cell = ws.cell(row=row_num+i, column=column_index_from_string('D'))
    #    cell.value = float(data.iloc[i]['changepercent'])

    #    cell = ws.cell(row=row_num+i, column=column_index_from_string('E'))
    #    cell.value = float(data.iloc[i]['volume']) / 100

    #    cell = ws.cell(row=row_num+i, column=column_index_from_string('F'))
    #    cell.value = float(data.iloc[i]['amount']) / 10000

    ## 设置列宽
    #column_widths = [12, 12, 10, 10, 12, 12]
    #for i in range(len(column_widths)):
    #    ws.column_dimensions[column_index_from_string('A')+i].width = column_widths[i]

    ## 保存Excel文件
    #wb.save(save_file)


if __name__ == '__main__':
    #obj_codes = ['000899','300913','300917','000791','600505','300175','000560','300542','300956','837046','301389','300947','300405','300637','300290']
    #obj_codes += ['300641','301012','300433','300905','002085','001696','300162','300159','002827','600101','300155','301252','301306','300502','300308','688525','301361','300843','301550','301387']
    
    obj_codes,_= readOneCsv('show_list')
    obj_codes = list(set(obj_codes))
    obj_codes = ['603900','000004','002583','002607','002423','600622','002693','000158','002456','600611','000062']

    #obj_codes = list(set(obj_codes))
    get_stock_data(obj_codes)
    #os.system('python getStockPrice.py 1')
    #adf = pd.read_csv("./dump/{}/stock_data_{}.csv".format(date[:4],date))
    #print(adf[['股票代码','股票名称','涨跌幅']])
    obj_codes =['sz','sh','cyb'] 
    get_stock_data(obj_codes)

    obj_codes = ['838971','833346']
    obj_codes,_= readOneCsv('show_bj')
    get_bj_stock(obj_codes)


